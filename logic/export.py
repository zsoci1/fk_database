import sqlite3
import os
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime

DB_PATH = "database/meals.db"


def extract_group_and_clean(address):
    if "#2" in address:
        return 2, address.replace("#2", "").strip()
    else:
        return 1, address.replace("#1", "").strip()

def export_delivery(date_str):
    os.makedirs("exports", exist_ok=True)

    # 1. Fetch data from DB
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute('''
        SELECT c.name, m.size, m.type_special,
               c.address1, c.address2, c.phone
        FROM meals m
        JOIN customers c ON m.customer_id = c.id
        WHERE m.date = ?
    ''', (date_str,))
    rows = cursor.fetchall()
    conn.close()

    # 2. Prepare data: group + clean + sort
    group1 = []
    group2 = []

    for row in rows:
        name, size, type_special, addr1, addr2, phone = row
        full_address = f"{addr1} {addr2}".strip()
        group, clean_address = extract_group_and_clean(full_address)
        values = [name, size, type_special, clean_address, phone, ""]

        if group == 1:
            group1.append(values)
        else:
            group2.append(values)

    group1.sort(key=lambda v: v[3])  # Sort by address
    group2.sort(key=lambda v: v[3])

    # 3. Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Delivery"

    # 4. Title row
    export_title = f"{date_str}"
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
    title_cell = ws.cell(row=1, column=1, value=export_title)
    title_cell.font = Font(name="Calibri", size=14, bold=True)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    # 5. Header row
    headers = ["Név", "Méret", "Étekezés", "Cím", "Tel.", "Megjegyzés"]
    ws.append(headers)
    header_font = Font(name="Calibri", size=12, bold=True)
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_num, value=header)
        cell.font = header_font
        cell.fill = PatternFill(start_color="D9D9D9", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # 6. Styles for data rows
    fill1 = PatternFill(start_color="FFFFFF", fill_type="solid")
    fill2 = PatternFill(start_color="F2F2F2", fill_type="solid")
    current_row = 3

    def write_group(title, data, row_start):
        nonlocal current_row
        ws.merge_cells(start_row=row_start, start_column=1, end_row=row_start, end_column=6)
        group_cell = ws.cell(row=row_start, column=1, value=title)
        group_cell.font = Font(size=12, bold=True)
        group_cell.alignment = Alignment(horizontal="left", vertical="center")
        current_row += 1

        for values in data:
            for j, value in enumerate(values, start=1):
                cell = ws.cell(row=current_row, column=j, value=value)
                cell.fill = fill1 if current_row % 2 == 0 else fill2
                cell.alignment = Alignment(wrap_text=True, vertical="center")
            current_row += 1

    # 7. Write both groups
    write_group("Délutáni címek", group1, current_row)
    current_row += 1  # Blank line
    write_group("Esti címek", group2, current_row)

    # 8. Center Size column (column B)
    for row in ws.iter_rows(min_row=3, min_col=2, max_col=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # 9. Row height tweak
    for row_idx in range(3, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 25

    # 10. Column widths (fit A4 width approx)
    ws.column_dimensions["A"].width = 18  # Name
    ws.column_dimensions["B"].width = 8   # Size
    ws.column_dimensions["C"].width = 60  # Type (wrap)
    ws.column_dimensions["D"].width = 28  # Address
    ws.column_dimensions["E"].width = 15  # Phone
    ws.column_dimensions["F"].width = 30  # Comment

    # 11. Save file
    filename = f"exports/delivery_{date_str}.xlsx"
    wb.save(filename)
    print(f"[✓] Delivery export completed → {filename}")

def parse_type_special(type_special_str):
    results = []

    items = [item.strip() for item in type_special_str.split(',')]
    for item in items:
        if ':' in item:
            base, special = item.split(':', 1)
            results.append((base.strip(), special.strip()))
        else:
            results.append((item.strip(), None))
    return results

def get_kitchen_summary(date_str):
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()

    cursor.execute('''
                   SELECT size, type_special
                   FROM meals
                   WHERE date = ?
                   ''', (date_str,))
    
    rows = cursor.fetchall()
    conn.close()

    summary = defaultdict(int)

    for size, type_special in rows:
        parsed = parse_type_special(type_special)
        for meal_type, special in parsed:
            summary[(meal_type, special, size)] += 1
    
    return summary

def export_kitchen(date_str):
    os.makedirs("exports", exist_ok=True)

    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()

    cursor.execute('''
        SELECT size, type_special FROM meals
        WHERE date = ?
    ''', (date_str,))
    rows = cursor.fetchall()
    conn.close()

    # Group and count meals: summary[size][(type, special)] = count
    summary = {}
    for size, type_special_str in rows:
        parsed = parse_type_special(type_special_str)
        if size not in summary:
            summary[size] = {}
        for meal_type, special in parsed:
            key = (meal_type, special)
            summary[size][key] = summary[size].get(key, 0) + 1

    size_order = ['S', 'M', 'L', 'XL']
    wb = Workbook()
    ws = wb.active
    ws.title = "Kitchen"

    # Title row
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)
    title_cell = ws.cell(row=1, column=1, value=f"{date_str}")
    title_cell.font = Font(size=14, bold=True)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    current_row = 3
    fill1 = PatternFill(start_color="FFFFFF", fill_type="solid")
    fill2 = PatternFill(start_color="F2F2F2", fill_type="solid")

    for size in size_order:
        if size not in summary:
            continue

        # Sub-header: Size: S
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=3)
        cell = ws.cell(row=current_row, column=1, value=f"Méret: {size}")
        cell.font = Font(bold=True, size=12)
        cell.alignment = Alignment(horizontal="left")
        current_row += 1

        # Column headers
        headers = ["Étkezés", "Spec.", "Mennyiség"]
        for col_num, header in enumerate(headers, start=1):
            cell = ws.cell(row=current_row, column=col_num, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D9D9D9", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        current_row += 1

        # Table content
        items = sorted(summary[size].items(), key=lambda x: (x[0][0], x[0][1] or ""))
        for i, ((meal_type, special), count) in enumerate(items):
            values = [meal_type, special or "", count]
            fill = fill1 if i % 2 == 0 else fill2
            for j, value in enumerate(values, start=1):
                cell = ws.cell(row=current_row, column=j, value=value)
                cell.fill = fill
                cell.alignment = Alignment(vertical="center")
            current_row += 1

        current_row += 1  # space between tables

    # Auto column widths
    for i, col in enumerate(ws.columns, start=1):
        max_length = 0
        col_letter = get_column_letter(i)
        for cell in col:
            try:
                max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2

    ws.column_dimensions["A"].width = 12

    filename = f"exports/kitchen_{date_str}.xlsx"
    wb.save(filename)
    print(f"[✓] Kitchen export completed → {filename}")

